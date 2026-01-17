import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from typing import Set, Tuple
import openpyxl

CELL_W = 80
CELL_H = 24
HEADER_H = 22
GRID = "#d0d0d0"
SEL = "#cde8ff"


# =====================================================
# Canvas-based Excel Table (cell-level selection)
# =====================================================
class ExcelCanvas(tk.Canvas):
    def __init__(self, parent, on_change):
        super().__init__(parent, bg="white", highlightthickness=0)
        self.on_change = on_change

        self.cells: Set[Tuple[int, int]] = set()
        self.anchor = None

        self.max_rows = 0
        self.max_cols = 0
        self._sync_target = None

        self.bind("<Button-1>", self._click)
        self.bind("<Control-Button-1>", self._ctrl)
        self.bind("<Shift-Button-1>", self._shift)
        self.bind("<MouseWheel>", self._wheel)

    # ---------- sync ----------
    def set_sync_target(self, other):
        self._sync_target = other

    def _wheel(self, e):
        delta = -1 if e.delta > 0 else 1
        f0, _ = self.yview()
        step = 0.02 * delta
        new_f = max(0.0, min(1.0, f0 + step))

        self.yview_moveto(new_f)
        if self._sync_target:
            self._sync_target.yview_moveto(new_f)

        return "break"

    # ---------- load ----------
    def load_sheet(self, path, sheet, max_rows, max_cols):
        self.delete("all")
        self.cells.clear()
        self.anchor = None

        self.max_rows = max_rows
        self.max_cols = max_cols

        if not path or not sheet:
            return

        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet not in wb.sheetnames:
            return

        ws = wb[sheet]
        rows = min(ws.max_row, max_rows)
        cols = min(ws.max_column, max_cols)

        # headers
        for c in range(max_cols):
            x = c * CELL_W
            self.create_rectangle(x, 0, x + CELL_W, HEADER_H,
                                  fill="#f0f0f0", outline=GRID)
            if c < cols:
                self.create_text(
                    x + CELL_W / 2,
                    HEADER_H / 2,
                    text=openpyxl.utils.get_column_letter(c + 1)
                )

        # cells
        for r in range(max_rows):
            for c in range(max_cols):
                x = c * CELL_W
                y = HEADER_H + r * CELL_H
                self.create_rectangle(
                    x, y, x + CELL_W, y + CELL_H,
                    outline=GRID,
                    tags=f"cell_{r}_{c}"
                )
                if r < rows and c < cols:
                    v = ws.cell(r + 1, c + 1).value
                    if v is not None:
                        self.create_text(
                            x + 4, y + CELL_H / 2,
                            anchor="w",
                            text=str(v),
                            tags=f"text_{r}_{c}"
                        )

        self.config(
            scrollregion=(0, 0,
                          max_cols * CELL_W,
                          HEADER_H + max_rows * CELL_H)
        )

    # ---------- selection helpers ----------
    def _cell_from_event(self, e):
        x = self.canvasx(e.x)
        y = self.canvasy(e.y)
        if y < HEADER_H:
            return None

        c = int(x // CELL_W)
        r = int((y - HEADER_H) // CELL_H)

        if r < 0 or c < 0 or r >= self.max_rows or c >= self.max_cols:
            return None
        return r, c

    def _clear_sel(self):
        for r, c in self.cells:
            self.itemconfig(f"cell_{r}_{c}", fill="white")

    def _apply_sel(self):
        for r, c in self.cells:
            self.itemconfig(f"cell_{r}_{c}", fill=SEL)

    # ---------- events ----------
    def _click(self, e):
        cell = self._cell_from_event(e)
        if not cell:
            return

        self._clear_sel()
        self.cells = {cell}
        self.anchor = cell
        self._apply_sel()
        self.on_change(self.cells)

    def _ctrl(self, e):
        cell = self._cell_from_event(e)
        if not cell:
            return

        if cell in self.cells:
            self.itemconfig(f"cell_{cell[0]}_{cell[1]}", fill="white")
            self.cells.remove(cell)
        else:
            self.cells.add(cell)
            self.itemconfig(f"cell_{cell[0]}_{cell[1]}", fill=SEL)

        self.anchor = cell
        self.on_change(self.cells)

    def _shift(self, e):
        if not self.anchor:
            return

        cell = self._cell_from_event(e)
        if not cell:
            return

        r1, c1 = self.anchor
        r2, c2 = cell

        self._clear_sel()
        self.cells = {
            (r, c)
            for r in range(min(r1, r2), max(r1, r2) + 1)
            for c in range(min(c1, c2), max(c1, c2) + 1)
        }
        self._apply_sel()
        self.on_change(self.cells)


# =====================================================
# Transfer CSV Builder Dialog
# =====================================================
class TransferCsvBuilderDialog(tk.Toplevel):

    HEADERS = [
        "source_file", "source_sheet", "source_cell",
        "source_row_offset", "source_col_offset",
        "destination_file", "destination_sheet", "destination_cell",
        "destination_row_offset", "destination_col_offset",
    ]

    def __init__(self, parent, ctx, logger, on_created):
        super().__init__(parent)

        self.ctx = ctx
        self.logger = logger
        self.on_created = on_created

        self.src_cells = set()
        self.dst_cells = set()

        self.max_rows = tk.IntVar(value=30)
        self.max_cols = tk.IntVar(value=10)

        self.title("転記CSV作成")
        self.geometry("1500x850")

        self._build()
        self._bind_table_keys()

    # ---------- focus restore ----------
    def _restore_focus(self):
        self.after(0, self.lift)
        self.after(0, self.focus_force)

    # =================================================
    def _build(self):
        root = ttk.Frame(self, padding=6)
        root.pack(fill="both", expand=True)
        root.columnconfigure(0, weight=1)
        root.columnconfigure(1, weight=1)
        root.rowconfigure(3, weight=1)

        # ---- 表示範囲 ----
        ctrl = ttk.LabelFrame(root, text="表示範囲")
        ctrl.grid(row=0, column=0, columnspan=2, sticky="we")

        ttk.Label(ctrl, text="行").pack(side="left")
        ttk.Spinbox(
            ctrl, from_=1, to=1000, width=6,
            textvariable=self.max_rows,
            command=self.reload_tables
        ).pack(side="left", padx=4)

        ttk.Label(ctrl, text="列").pack(side="left", padx=(12, 0))
        ttk.Spinbox(
            ctrl, from_=1, to=100, width=6,
            textvariable=self.max_cols,
            command=self.reload_tables
        ).pack(side="left", padx=4)

        # ---- 候補 ----
        cand = ttk.LabelFrame(root, text="マッピング候補")
        cand.grid(row=1, column=0, columnspan=2, sticky="we")
        self.cand_var = tk.StringVar()
        ttk.Entry(cand, textvariable=self.cand_var,
                  state="readonly").pack(fill="x", padx=4, pady=4)

        # ---- 確定テーブル ----
        conf = ttk.LabelFrame(root, text="確定済み（CSV内容）")
        conf.grid(row=2, column=0, columnspan=2, sticky="nsew")
        conf.rowconfigure(0, weight=1)
        conf.columnconfigure(0, weight=1)

        tv = ttk.Frame(conf)
        tv.grid(row=0, column=0, sticky="nsew")

        self.table = ttk.Treeview(tv, columns=self.HEADERS, show="headings", selectmode="extended")
        for h in self.HEADERS:
            self.table.heading(h, text=h)
            self.table.column(h, width=140, anchor="center")

        vsb = ttk.Scrollbar(tv, orient="vertical", command=self.table.yview)
        hsb = ttk.Scrollbar(tv, orient="horizontal", command=self.table.xview)
        self.table.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.table.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        tv.rowconfigure(0, weight=1)
        tv.columnconfigure(0, weight=1)

        btns = ttk.Frame(conf)
        btns.grid(row=1, column=0, sticky="we", pady=4)
        ttk.Button(btns, text="確定", command=self.confirm).pack(side="left")
        ttk.Button(btns, text="削除", command=self.remove).pack(side="left", padx=6)
        ttk.Button(btns, text="CSVを作成...", command=self.create_csv).pack(side="right")

        # ---- Excel 表 ----
        paned = ttk.Panedwindow(root, orient="horizontal")
        paned.grid(row=3, column=0, columnspan=2, sticky="nsew")

        self._build_side(paned, True)
        self._build_side(paned, False)

        self.src_canvas.set_sync_target(self.dst_canvas)
        self.dst_canvas.set_sync_target(self.src_canvas)

    # =================================================
    # ★ 追加：確定テーブルのキー操作
    # =================================================
    def _bind_table_keys(self):
        self.table.bind("<Control-a>", self._select_all_rows)
        self.table.bind("<Control-A>", self._select_all_rows)
        self.table.bind("<Delete>", self._delete_selected_rows)
        self.table.bind("<BackSpace>", self._delete_selected_rows)

    def _select_all_rows(self, event=None):
        if self.focus_get() is not self.table:
            return "break"
        items = self.table.get_children()
        if items:
            self.table.selection_set(items)
        return "break"

    def _delete_selected_rows(self, event=None):
        if self.focus_get() is not self.table:
            return "break"
        for iid in self.table.selection():
            self.table.delete(iid)
        return "break"

    # =================================================
    def _build_side(self, paned, is_src):
        frm = ttk.Frame(paned)
        paned.add(frm, weight=1)

        title = "ファイルA（転記元）" if is_src else "ファイルB（転記先）"
        lf = ttk.LabelFrame(frm, text=title)
        lf.pack(fill="x")

        ttk.Label(lf, text="ファイル").grid(row=0, column=0)
        ent = tk.Entry(lf)
        ent.grid(row=0, column=1, sticky="we")
        ttk.Button(
            lf, text="参照...",
            command=lambda e=ent, s=is_src: self.choose_file(e, s)
        ).grid(row=0, column=2)

        ttk.Label(lf, text="シート").grid(row=1, column=0)
        cmb = ttk.Combobox(lf, state="readonly")
        cmb.grid(row=1, column=1, sticky="w")

        canvas = ExcelCanvas(
            frm,
            self.on_src_change if is_src else self.on_dst_change
        )
        canvas.pack(fill="both", expand=True)

        if is_src:
            self.src_file, self.src_sheet, self.src_canvas = ent, cmb, canvas
        else:
            self.dst_file, self.dst_sheet, self.dst_canvas = ent, cmb, canvas

        cmb.bind(
            "<<ComboboxSelected>>",
            lambda _e, e=ent, c=cmb, cv=canvas:
            cv.load_sheet(
                e.get(), c.get(),
                self.max_rows.get(), self.max_cols.get()
            )
        )

    # =================================================
    def reload_tables(self):
        if self.src_file.get() and self.src_sheet.get():
            self.src_canvas.load_sheet(
                self.src_file.get(),
                self.src_sheet.get(),
                self.max_rows.get(),
                self.max_cols.get()
            )
        if self.dst_file.get() and self.dst_sheet.get():
            self.dst_canvas.load_sheet(
                self.dst_file.get(),
                self.dst_sheet.get(),
                self.max_rows.get(),
                self.max_cols.get()
            )

    def on_src_change(self, cells):
        self.src_cells = cells
        self.refresh_candidate()

    def on_dst_change(self, cells):
        self.dst_cells = cells
        self.refresh_candidate()

    def refresh_candidate(self):
        sc = len(self.src_cells)
        dc = len(self.dst_cells)

        if sc:
            names = self._cell_name(self.src_cells)
            s_txt = f"{names[0]}〜{names[-1]} ({sc})"
        else:
            s_txt = "(未選択)"

        if dc:
            names = self._cell_name(self.dst_cells)
            d_txt = f"{names[0]}〜{names[-1]} ({dc})"
        else:
            d_txt = "(未選択)"

        self.cand_var.set(f"A: {s_txt}  →  B: {d_txt}")

    def _cell_name(self, cells):
        return [
            f"{openpyxl.utils.get_column_letter(c + 1)}{r + 1}"
            for r, c in sorted(cells)
        ]

    def confirm(self):
        if not self.src_cells or not self.dst_cells:
            return

        src = self._cell_name(self.src_cells)
        dst = self._cell_name(self.dst_cells)

        if len(src) == len(dst):
            pairs = zip(src, dst)
        elif len(src) == 1:
            pairs = [(src[0], d) for d in dst]
        elif len(dst) == 1:
            pairs = [(s, dst[0]) for s in src]
        else:
            messagebox.showerror("転記", "セル数が一致しません")
            self._restore_focus()
            return

        for s, d in pairs:
            self.table.insert("", "end", values=[
                self.src_file.get(), self.src_sheet.get(), s, 0, 0,
                self.dst_file.get(), self.dst_sheet.get(), d, 0, 0
            ])

        self.src_cells.clear()
        self.dst_cells.clear()
        self.refresh_candidate()

    def remove(self):
        for iid in self.table.selection():
            self.table.delete(iid)

    def create_csv(self):
        if not self.table.get_children():
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv")]
        )
        self._restore_focus()

        if not path:
            return

        lines = [",".join(self.HEADERS)]
        for iid in self.table.get_children():
            lines.append(",".join(map(str, self.table.item(iid)["values"])))


        Path(path).write_text("\n".join(lines) + "\n", encoding="utf-8-sig")

        if self.on_created:
            self.on_created(path)
        self.destroy()

    def choose_file(self, entry, is_src):
        p = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsx;*.xlsm;*.xlsb;*.xls")]
        )
        self._restore_focus()

        if not p:
            return

        entry.delete(0, tk.END)
        entry.insert(0, p)

        wb = openpyxl.load_workbook(p, read_only=True)
        sheets = wb.sheetnames
        cmb = self.src_sheet if is_src else self.dst_sheet
        cmb["values"] = sheets
        if sheets:
            cmb.set(sheets[0])
            self.reload_tables()
