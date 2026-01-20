# ============================================
# excel_transfer/services/excel_view_service.py
# ============================================
from __future__ import annotations

import os
import time
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import openpyxl


@dataclass
class ExcelViewport:
    top_row: int = 1
    left_col: int = 1
    rows: int = 30
    cols: int = 10


class ExcelViewService:
    """
    Excel 読み取り（プレビュー用）
    - openpyxl(read_only=True, data_only=True)
    - 65535 / 1048576 行でも「表示範囲のみ」読み出し
    - 複数ブック(複数ファイル)保持、book/sheet を dropdown で切替

    ※転記処理 (xlwings/COM) とは分離している
    """

    def __init__(self, logger=None):
        self.logger = logger

        self._books: Dict[str, openpyxl.Workbook] = {}
        self._book_paths: List[str] = []

        self._current_book_path: str = ""
        self._current_sheet_name: str = ""

        self._ws = None
        self._max_row: int = 0
        self._max_col: int = 0

        # セル値キャッシュ（表示範囲用途）
        self._cache: Dict[Tuple[int, int], str] = {}
        self._cache_ts: float = 0.0
        self._cache_ttl_sec: float = 0.5  # 連続スクロールの負荷軽減

    # -------------------------
    # logging
    # -------------------------
    def _log(self, level: str, msg: str) -> None:
        try:
            if self.logger:
                fn = getattr(self.logger, level.lower(), None)
                if fn:
                    fn(msg)
                else:
                    self.logger.info(msg)
        except Exception:
            pass

    # -------------------------
    # book management
    # -------------------------
    def add_books(self, paths: List[str]) -> None:
        add = []
        for p in paths:
            if not p:
                continue
            p = os.path.abspath(p)
            if p in self._books:
                continue
            if not os.path.exists(p):
                self._log("error", f"[ExcelViewService] missing file: {p}")
                continue
            add.append(p)

        for p in add:
            try:
                wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                self._books[p] = wb
                self._book_paths.append(p)
                self._log("info", f"[ExcelViewService] book loaded: {p}")
            except Exception as e:
                self._log("error", f"[ExcelViewService] load failed: {p} err={e}")

        if not self._current_book_path and self._book_paths:
            self.select_book(self._book_paths[0])

    def get_book_paths(self) -> List[str]:
        return list(self._book_paths)

    def select_book(self, path: str) -> None:
        path = os.path.abspath(path) if path else ""
        if not path or path not in self._books:
            self._log("error", f"[ExcelViewService] select_book invalid: {path}")
            self._current_book_path = ""
            self._current_sheet_name = ""
            self._ws = None
            self._max_row = 0
            self._max_col = 0
            self._cache.clear()
            return

        self._current_book_path = path
        wb = self._books[path]
        sheets = wb.sheetnames
        self._log("info", f"[ExcelViewService] select book: {path} sheets={len(sheets)}")

        if sheets:
            self.select_sheet(sheets[0])
        else:
            self._current_sheet_name = ""
            self._ws = None
            self._max_row = 0
            self._max_col = 0
            self._cache.clear()

    def get_sheet_names(self) -> List[str]:
        if not self._current_book_path:
            return []
        wb = self._books.get(self._current_book_path)
        if not wb:
            return []
        return list(wb.sheetnames)

    def select_sheet(self, sheet_name: str) -> None:
        if not self._current_book_path:
            return
        wb = self._books.get(self._current_book_path)
        if not wb:
            return
        if not sheet_name or sheet_name not in wb.sheetnames:
            self._log("error", f"[ExcelViewService] select_sheet invalid: {sheet_name}")
            return

        self._current_sheet_name = sheet_name
        self._ws = wb[sheet_name]
        self._max_row = int(getattr(self._ws, "max_row", 0) or 0)
        self._max_col = int(getattr(self._ws, "max_column", 0) or 0)
        self._cache.clear()
        self._cache_ts = 0.0
        self._log("info", f"[ExcelViewService] select sheet: {sheet_name} max=({self._max_row},{self._max_col})")

    def get_current_book_path(self) -> str:
        return self._current_book_path

    def get_current_sheet_name(self) -> str:
        return self._current_sheet_name

    def get_sheet_size(self) -> Tuple[int, int]:
        return self._max_row, self._max_col

    # -------------------------
    # value access (viewport)
    # -------------------------
    def clear_cache(self) -> None:
        self._cache.clear()
        self._cache_ts = 0.0

    def _cache_valid(self) -> bool:
        if not self._cache:
            return False
        return (time.time() - self._cache_ts) <= self._cache_ttl_sec

    def prime_viewport_cache(self, vp: ExcelViewport) -> None:
        """
        表示領域だけキャッシュ（スクロール中のチラつき・遅延を抑える）
        """
        if not self._ws:
            return

        # 直近キャッシュが生きているなら無駄に更新しない
        if self._cache_valid():
            return

        self._cache.clear()

        top = max(1, int(vp.top_row))
        left = max(1, int(vp.left_col))
        bottom = max(top, top + int(vp.rows) - 1)
        right = max(left, left + int(vp.cols) - 1)

        # シート最大の範囲に丸める
        if self._max_row > 0:
            bottom = min(bottom, self._max_row)
        if self._max_col > 0:
            right = min(right, self._max_col)

        # openpyxl は range でまとめて取ると速い（read_onlyの制約あり）
        # ただし巨大範囲は禁止。vpに依存。
        try:
            for r in range(top, bottom + 1):
                for c in range(left, right + 1):
                    v = self._ws.cell(row=r, column=c).value
                    if v is None:
                        s = ""
                    else:
                        s = str(v)
                    self._cache[(r, c)] = s
            self._cache_ts = time.time()
        except Exception as e:
            self._log("error", f"[ExcelViewService] prime cache failed: {e}")
            self._cache.clear()
            self._cache_ts = 0.0

    def get_cell_text(self, row: int, col: int) -> str:
        if not self._ws or row <= 0 or col <= 0:
            return ""
        if (row, col) in self._cache:
            return self._cache[(row, col)]

        try:
            v = self._ws.cell(row=row, column=col).value
            return "" if v is None else str(v)
        except Exception:
            return ""

    # -------------------------
    # close
    # -------------------------
    def close(self) -> None:
        for p, wb in list(self._books.items()):
            try:
                wb.close()
            except Exception:
                pass
        self._books.clear()
        self._book_paths.clear()
        self._current_book_path = ""
        self._current_sheet_name = ""
        self._ws = None
        self._max_row = 0
        self._max_col = 0
        self._cache.clear()
        self._cache_ts = 0.0
