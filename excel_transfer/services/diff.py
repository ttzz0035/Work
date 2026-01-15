# excel_transfer/services/diff.py
from __future__ import annotations

import os
import sys
import json
import shutil
from typing import Tuple, Dict, Any, List

# =====================================================
# project root 探索
# =====================================================
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))

def _resolve_project_root(start: str) -> str:
    cur = start
    for _ in range(10):
        cand = os.path.join(cur, "models", "dto.py")
        if os.path.isfile(cand):
            return cur
        parent = os.path.dirname(cur)
        if parent == cur:
            break
        cur = parent
    raise RuntimeError("project root not found")

_PROJECT_ROOT = _resolve_project_root(_THIS_DIR)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

# =====================================================
# imports
# =====================================================
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

from models.dto import DiffRequest, LogFn

# =====================================================
# Excel Diff Service
# =====================================================
class ExcelDiffService:
    def __init__(self, req: DiffRequest, logger, append_log: LogFn):
        self.req = req
        self.logger = logger
        self.append_log = append_log

        self.diff_cells: List[Dict[str, Any]] = []
        self.diff_shapes: List[Dict[str, Any]] = []
        self.rows_added: List[Dict[str, Any]] = []
        self.rows_deleted: List[Dict[str, Any]] = []

        self.out_file: str = ""
        self.json_file: str = ""

    # -------------------------------------------------
    def run(self) -> str:
        self._log("=== Diff開始 ===")

        if not os.path.exists(self.req.file_a) or not os.path.exists(self.req.file_b):
            raise ValueError("比較ファイルが存在しません")

        base = os.path.splitext(self.req.file_b)[0]
        self.out_file = base + "_DIFF.xlsx"
        self.json_file = base + "_DIFF.json"

        shutil.copyfile(self.req.file_b, self.out_file)
        self._log(f"[INFO] 出力Excel: {self.out_file}")

        self._diff_excel()
        self._mark_cells_red()
        self._mark_shapes_red()
        self._write_json()

        self._log(f"[INFO] 出力JSON: {self.json_file}")
        self._log("=== Diff完了 ===")
        return self.out_file

    # -------------------------------------------------
    def _diff_excel(self) -> None:
        app_a = app_b = None
        book_a = book_b = None

        try:
            app_a = xw.App(visible=False, add_book=False)
            app_b = xw.App(visible=False, add_book=False)

            book_a = app_a.books.open(self.req.file_a, read_only=True)
            book_b = app_b.books.open(self.req.file_b, read_only=True)

            sht_a = book_a.sheets[0]
            sht_b = book_b.sheets[0]

            a_dict = self._read_sheet_to_dict(sht_a)
            b_dict = self._read_sheet_to_dict(sht_b)

            a_keys = set(a_dict.keys())
            b_keys = set(b_dict.keys())

            for k in a_keys - b_keys:
                self.rows_deleted.append({"key": list(k)})
                self._log(f"[DEL] {k}")

            for k in b_keys - a_keys:
                self.rows_added.append({"key": list(k)})
                self._log(f"[ADD] {k}")

            for k in a_keys & b_keys:
                ra = a_dict[k]
                rb = b_dict[k]
                for col in ra:
                    va = ra[col]["value"]
                    vb = rb[col]["value"]
                    if va != vb:
                        self.diff_cells.append({
                            "key": list(k),
                            "column": col,
                            "a": va,
                            "b": vb,
                            "row": rb[col]["row"],
                            "col": rb[col]["col"],
                        })
                        self._log(f"[MOD] {k} {col}: {va} -> {vb}")

            if self.req.compare_shapes:
                sa = self._read_shapes(sht_a)
                sb = self._read_shapes(sht_b)

                for name in sb.keys() - sa.keys():
                    self.diff_shapes.append({"name": name, "type": "ADD"})
                for name in sa.keys() - sb.keys():
                    self.diff_shapes.append({"name": name, "type": "DEL"})
                for name in sa.keys() & sb.keys():
                    if sa[name] != sb[name]:
                        self.diff_shapes.append({"name": name, "type": "MOD"})

        finally:
            for book in (book_a, book_b):
                try:
                    if book:
                        book.close()
                except Exception:
                    pass
            for app in (app_a, app_b):
                try:
                    if app:
                        app.kill()
                except Exception:
                    pass

    # -------------------------------------------------
    def _read_sheet_to_dict(self, sht: xw.Sheet) -> Dict[Tuple, Dict[str, Any]]:
        vals = sht.used_range.value
        if not vals:
            return {}

        headers = vals[0]
        col_index = {str(h): i for i, h in enumerate(headers)}
        rows = vals[1:]

        out: Dict[Tuple, Dict[str, Any]] = {}

        for ridx, row in enumerate(rows, start=2):
            key = (
                tuple(row[col_index[k]] for k in self.req.key_cols)
                if self.req.key_cols else (ridx,)
            )
            record: Dict[str, Any] = {}
            for h, cidx in col_index.items():
                cell = sht.range((ridx, cidx + 1))
                v = cell.formula if self.req.compare_formula else cell.value
                record[h] = {"value": v, "row": ridx, "col": cidx + 1}
            out[key] = record

        return out

    def _read_shapes(self, sht: xw.Sheet) -> Dict[str, Dict[str, Any]]:
        out = {}
        for shp in sht.api.Shapes:
            try:
                out[str(shp.Name)] = {
                    "top": shp.Top,
                    "left": shp.Left,
                    "width": shp.Width,
                    "height": shp.Height,
                }
            except Exception:
                pass
        return out

    # -------------------------------------------------
    def _mark_cells_red(self) -> None:
        if not self.diff_cells:
            return

        wb = load_workbook(self.out_file)
        ws = wb.active

        red = PatternFill(start_color="FFFF6666", end_color="FFFF6666", fill_type="solid")
        border = Border(
            left=Side(style="thin", color="FF0000"),
            right=Side(style="thin", color="FF0000"),
            top=Side(style="thin", color="FF0000"),
            bottom=Side(style="thin", color="FF0000"),
        )

        for d in self.diff_cells:
            cell = ws.cell(row=int(d["row"]), column=int(d["col"]))
            cell.fill = red
            cell.border = border

        wb.save(self.out_file)

    def _mark_shapes_red(self) -> None:
        if not self.diff_shapes:
            return

        app = xw.App(visible=False, add_book=False)
        try:
            book = app.books.open(self.out_file)
            sht = book.sheets[0]
            for shp in sht.api.Shapes:
                for d in self.diff_shapes:
                    if shp.Name == d["name"]:
                        try:
                            shp.Line.ForeColor.RGB = 255
                        except Exception:
                            pass
            book.save()
            book.close()
        finally:
            app.kill()

    # -------------------------------------------------
    def _write_json(self) -> None:
        data = {
            "meta": {
                "file_a": self.req.file_a,
                "file_b": self.req.file_b,
                "compare_formula": self.req.compare_formula,
                "compare_shapes": self.req.compare_shapes,
                "key_cols": self.req.key_cols,
            },
            "summary": {
                "added_rows": len(self.rows_added),
                "deleted_rows": len(self.rows_deleted),
                "modified_cells": len(self.diff_cells),
                "shape_diffs": len(self.diff_shapes),
            },
            "rows": {
                "added": self.rows_added,
                "deleted": self.rows_deleted,
            },
            "cells": self.diff_cells,
            "shapes": self.diff_shapes,
        }

        with open(self.json_file, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    # -------------------------------------------------
    def _log(self, msg: str):
        if self.logger:
            try:
                self.logger.info(msg)
            except Exception:
                pass
        if self.append_log:
            self.append_log(msg)


# =====================================================
def run_diff(req: DiffRequest, ctx, logger, append_log: LogFn) -> str:
    return ExcelDiffService(req, logger, append_log).run()
