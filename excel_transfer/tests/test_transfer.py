# tests/test_transfer.py
import csv
from pathlib import Path
import openpyxl
import xlwings as xw
from services.transfer import run_transfer_from_csvs
from models.dto import TransferRequest

def test_transfer_basic(ctx, tmp_path):
    # 準備: ソース/宛先ブックと転記CSV
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"

    # ソース: Sheet1!B2= "Hello", C2= 123
    from tests.conftest import wb_write
    wb_write(src, {"Sheet1": [["H1","H2","H3"], ["row1","Hello",123]]})
    wb_write(dst, {"Out": [["X1","X2","X3"], ["", "", ""]]})

    csv_path = tmp_path / "map.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "source_file","source_sheet","source_cell","source_row_offset","source_col_offset",
            "destination_file","destination_sheet","destination_cell","destination_row_offset","destination_col_offset"
        ])
        # B2 -> B2、 C2 -> C2（Outシート）
        w.writerow([src.name,"Sheet1","B2",0,0,dst.name,"Out","B2",0,0])
        w.writerow([src.name,"Sheet1","C2",0,0,dst.name,"Out","C2",0,0])

    # 実行
    req = TransferRequest(csv_paths=[str(csv_path)])
    note = run_transfer_from_csvs(req, ctx, logger=None, append_log=lambda *_: None)
    assert Path(note).name == "map.csv"

    # 検証: 宛先が更新されているか
    wb = openpyxl.load_workbook(dst)
    ws = wb["Out"]
    assert ws["B2"].value == "Hello"
    assert ws["C2"].value == 123
