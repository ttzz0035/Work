# tests/test_diff.py
from pathlib import Path
import openpyxl
from services.diff import run_diff
from models.dto import DiffRequest
from tests.conftest import wb_write

def _read_summary_counts(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Summary"]
    # Summaryは A7:D? に ["Sheet","Changed","Added","Deleted"]
    # A8〜の合計行はない想定。ここでは1行目分を拾う。
    headers = [ws.cell(7,1).value, ws.cell(7,2).value, ws.cell(7,3).value, ws.cell(7,4).value]
    assert headers == ["Sheet","Changed","Added","Deleted"]
    # 次行
    name = ws.cell(8,1).value
    changed = ws.cell(8,2).value or 0
    added = ws.cell(8,3).value or 0
    deleted = ws.cell(8,4).value or 0
    return name, int(changed), int(added), int(deleted)

def test_diff_position_based(ctx, tmp_path):
    a = tmp_path / "A.xlsx"
    b = tmp_path / "B.xlsx"
    # ヘッダ1行 + 2行のデータ、B2が異なる
    wb_write(a, {"S": [["col1","col2"], ["r1","x"], ["r2","y"]]})
    wb_write(b, {"S": [["col1","col2"], ["r1","X"], ["r2","y"]]})

    out = run_diff(
        DiffRequest(file_a=str(a), file_b=str(b), key_cols=[], compare_formula=False),
        ctx, logger=None, append_log=lambda *_: None
    )
    assert Path(out).name == "diff_report.xlsx"

    name, changed, added, deleted = _read_summary_counts(Path(out))
    assert name == "S"
    assert changed >= 1  # 少なくとも1セル差分
    assert added == 0 and deleted == 0

def test_diff_key_based(ctx, tmp_path):
    a = tmp_path / "A2.xlsx"
    b = tmp_path / "B2.xlsx"
    # キーは "ID"
    wb_write(a, {"S": [
        ["ID","Name","Val"],
        ["1","Alice","100"],
        ["2","Bob","200"]
    ]})
    wb_write(b, {"S": [
        ["ID","Name","Val"],
        ["1","Alice","150"],  # 変更
        ["3","Carol","300"]   # 追加
    ]})

    out = run_diff(
        DiffRequest(file_a=str(a), file_b=str(b), key_cols=["ID"], compare_formula=False),
        ctx, logger=None, append_log=lambda *_: None
    )
    assert Path(out).name == "diff_report.xlsx"

    # Summary: Changed=1（ID=1のVal）、Added=1（ID=3）、Deleted=1（ID=2）
    name, changed, added, deleted = _read_summary_counts(Path(out))
    assert name == "S"
    assert changed == 1
    assert added == 1
    assert deleted == 1

    # 詳細シートの存在も軽く確認
    wb = openpyxl.load_workbook(out)
    assert any(s.title.startswith("Changed_") for s in wb.worksheets)
    assert any(s.title.startswith("Added_") for s in wb.worksheets)
    assert any(s.title.startswith("Deleted_") for s in wb.worksheets)
